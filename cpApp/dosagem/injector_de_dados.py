import pandas as pd
from openpyxl import load_workbook

def motor_horizontal(arquivo_nome, lista_medicamentos):
    try:
        # 1. Abre o ficheiro preservando estilos e cabeçalhos
        wb = load_workbook(arquivo_nome)
        ws = wb['oms']

        # 2. Localizar a primeira linha vaga a partir da 8
        linha_vaga = 8
        while ws.cell(row=linha_vaga, column=1).value is not None:
            linha_vaga += 1
        
        print(f"Começando a inserir dados a partir da linha: {linha_vaga}")

        # 3. Inserção cega (Esquerda para Direita)
        # Cada 'med' deve ter exatamente 9 itens para bater com as colunas A até I
        for med in lista_medicamentos:
            for col_idx, valor in enumerate(med, start=1):
                ws.cell(row=linha_vaga, column=col_idx).value = valor
            linha_vaga += 1

        # 4. Salvar no próprio arquivo
        wb.save(arquivo_nome)
        print(f"Sucesso! {len(lista_medicamentos)} linhas adicionadas.  Mpindi TecMed")

    except Exception as e:
        print(f"Erro no Pydroid: {e}")



# ============================================================================
# BLOCO 15: 37 OFTALMOLÓGICOS, SUPLEMENTOS E CORREÇÕES - OMS 2025 / MSF 2024
# ============================================================================

novos_dados = [
    
    # =========================================================================
    # OFTALMOLÓGICOS INJECTÁVEIS
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 1. ACETAZOLAMIDA (Glaucoma agudo, Edema cerebral)
    # -------------------------------------------------------------------------
    ["Acetazolamida", "Glaucoma", "", "IV", "dosagem(250, 500, 500, mg)", "500 mg/5ml|100", "6, 8", "%[{#d/#c/#i}] ml%#Diluir em 10 ml ABD e infundir em 5 min#@@Acidose metabólica! Hipocalemia!@", "OMS 2025. MSF 2024. Glaucoma agudo."],
    
    ["Acetazolamida", "Edema cerebral", "", "IV", "dosagem(250, 500, 250, mg)", "500 mg/5ml|100", "6, 8", "%[{#d/#c/#i}] ml%#Diluir em 10 ml ABD e infundir em 5 min", "OMS 2025."],

    # -------------------------------------------------------------------------
    # 2. BEVACIZUMAB (Anti-VEGF) (Degeneração macular, Retinopatia)
    # -------------------------------------------------------------------------
    ["Bevacizumab", "", "", "", "dosagem(1.25, 1.25, 1.25, mg)", "100 mg/4ml|25", "", "%[{#d/#c}] ml%#Intravítreo. Técnica estéril.#@@Endoftalmite!@", "OMS 2025. DMRI, retinopatia diabética."],

    # -------------------------------------------------------------------------
    # 3. RANIBIZUMAB (Anti-VEGF)
    # -------------------------------------------------------------------------
    ["Ranibizumab", "", "", "", "dosagem(0.5, 0.5, 0.5, mg)", "10 mg/ml|10", "", "%[{#d/#c}] ml%#Intravítreo. Técnica estéril.", "OMS 2025. DMRI."],

    # =========================================================================
    # SUPLEMENTOS E CORREÇÕES
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 4. ALBUMINA HUMANA 20% (Hipovolemia, Queimaduras, Paracentese)
    # -------------------------------------------------------------------------
    ["Albumina Humana 20%", "Hipovolemia", "", "", "dosagem(0.5, 1, 1, g/kg)", "200 mg/ml|200", "", "%[{#p*#d/#c}] ml%#Infundir em 30-60 min. Sem diluir.#@@Anafilaxia rara!@", "OMS 2025. MSF 2024."],
    
    ["Albumina Humana 20%", "Queimaduras", "", "", "dosagem(0.5, 1, 1, g/kg)", "200 mg/ml|200", "8, 12", "%[{#p*#d/#c/#i}] ml%#Infundir em 30-60 min.", "OMS 2025. >30% SCQ."],
    
    ["Albumina Humana 20%", "Paracentese", "", "", "dosagem(6, 8, 8, g/L aspirado)", "200 mg/ml|200", "", "%[{#d/#co}] ml%#Infundir em 30-60 min. Após paracentese >5L.", "OMS 2025. Prevenção disfunção circulatória."],

    # -------------------------------------------------------------------------
    # 5. FOSFATO DE POTÁSSIO (Hipofosfatemia)
    # -------------------------------------------------------------------------
    ["Fosfato de Potássio", "", "", "", "dosagem(0.08, 0.16, 0.16, mmol/kg)", "1 mmol/ml|1", "6, 12", "%[{#p*#d/#c/#i}] ml%#Diluir em 250 ml SF e infundir em 4-6h#@@Hiperfosfatemia! Hipocalcemia!@", "OMS 2025. Hipofosfatemia grave."],

    # -------------------------------------------------------------------------
    # 6. SULFATO DE ZINCO (Deficiência - NPT)
    # -------------------------------------------------------------------------
    ["Sulfato de Zinco", "", "", "", "dosagem(2, 5, 5, mg)", "5 mg/ml|5", "24", "%[{#d/#c}] ml%#Adicionar à solução NPT.", "OMS 2025. Nutrição parenteral."],

    # =========================================================================
    # IMUNOMODULADORES E BIOLÓGICOS
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 7. HIDROXICLOROQUINA (Malária, LES - off label IV)
    # -------------------------------------------------------------------------
    # NOTA: Hidroxicloroquina é primariamente oral. Pular.

    # -------------------------------------------------------------------------
    # 8. METOTREXATO (Artrite reumatoide, Gravidez ectópica)
    # -------------------------------------------------------------------------
    ["Metotrexato", "Artrite", "", "IM", "dosagem(7.5, 25, 15, mg)", "25 mg/ml|25", "7", "%[{#d/#c}] ml%#IM profunda. Dose semanal.#@@Mielotoxicidade! Ácido Fólico!@", "OMS 2025. AR, psoríase."],
    
    ["Metotrexato", "Gravidez ectópica", "", "IM", "dosagem(50, 50, 50, mg/m2)", "25 mg/ml|25", "", "%[{#d/#c}] ml%#IM profunda. Dose única.#@@Confirmar β-hCG <5000!@", "OMS 2025. Gravidez ectópica."],

    # -------------------------------------------------------------------------
    # 9. CICLOFOSFAMIDA (Pulso imunossupressor)
    # -------------------------------------------------------------------------
    ["Ciclofosfamida", "", "", "", "dosagem(500, 1000, 750, mg)", "500 mg/25ml|20", "12, 24", "%[{#d/#c/#i}] ml%#Diluir em 100 ml SF e infundir em 30 min#@@Cistite hemorrágica! Hidratar! Mesna!@", "OMS 2025. LES, vasculites."],

    # =========================================================================
    # ANTIESPASMÓDICOS E MOTILIDADE
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 10. ESCOPOLAMINA (Butilbrometo) (Cólica, Pré-medicação)
    # -------------------------------------------------------------------------
    ["Escopolamina", "", "adulta", "IV", "peso(40, 120); dosagem(20, 40, 20, mg)", "20 mg/ml|20", "6, 8", "%[{#d/#c/#i}] ml%#IV directo lento em 2-5 min.", "OMS 2025. Antiespasmódico."],
    
    ["Escopolamina", "", "adulta", "IM", "peso(40, 120); dosagem(20, 40, 20, mg)", "20 mg/ml|20", "6, 8", "%[{#d/#c/#i}] ml%#IM profunda.", "OMS 2025."],
    
    ["Escopolamina", "", "pediatrica", "IV", "peso(3, 40); dosagem(0.3, 0.6, 0.5, mg/kg)", "20 mg/ml|20", "6, 8", "%[{#p*#d/#c/#i}] ml%#IV directo lento.", "OMS 2025."],

    # =========================================================================
    # BRONCODILATADORES ADICIONAIS
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 11. BROMETO DE IPRATRÓPIO (Asma - nebulização)
    # -------------------------------------------------------------------------
    # NOTA: Ipratrópio é nebulizado. Pular para injectáveis.
    
    # -------------------------------------------------------------------------
    # 12. TERBUTALINA (Asma grave, Tocólise)
    # -------------------------------------------------------------------------
    ["Terbutalina", "Asma", "adulta", "SC", "peso(40, 120); dosagem(0.25, 0.5, 0.25, mg)", "0.5 mg/ml|0.5", "4, 6", "%[{#d/#c/#i}] ml%#SC profunda. Máx 4 doses.#@@Taquicardia!@", "OMS 2025. Asma aguda."],
    
    ["Terbutalina", "Asma", "adulta", "IV", "peso(40, 120); dosagem(0.25, 0.5, 0.25, mg)", "0.5 mg/ml|0.5", "4, 6", "%[{#d/#c/#i}] ml%#Diluir em 10 ml SF e infundir em 5 min", "OMS 2025."],

    # -------------------------------------------------------------------------
    # 13. SULFATO DE MAGNÉSIO (Asma grave - adjuvante)
    # -------------------------------------------------------------------------
    ["Sulfato de Magnésio", "Asma", "adulta", "", "peso(40, 120); dosagem(40, 50, 40, mg/kg)", "500 mg/ml|500", "", "%[{#p*#d/#c}] ml%#Diluir em 100 ml SF e infundir em 30 min#@@Monitorizar reflexo patelar! FR!@", "OMS 2025. Asma aguda grave."],

    # =========================================================================
    # EMERGÊNCIAS ENDÓCRINAS
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 14. SOLUÇÃO DE LUGOL (Iodo) (Crise tireotóxica)
    # -------------------------------------------------------------------------
    ["Solução de Lugol", "", "", "IV", "dosagem(5, 10, 10, gotas)", "50 mg/ml|50", "6, 8", "%[{#d/#c/#i}] ml%#Diluir em 100 ml SF e infundir em 30 min#@@Iniciar após tionamidas!@", "OMS 2025. Crise tireotóxica."],

    # -------------------------------------------------------------------------
    # 15. METIMAZOL (Crise tireotóxica)
    # -------------------------------------------------------------------------
    ["Metimazol", "", "", "IV", "dosagem(20, 40, 20, mg)", "10 mg/ml|10", "6, 8", "%[{#d/#c/#i}] ml%#Diluir em 50 ml SF e infundir em 15 min#@@Agranulocitose!@", "OMS 2025. Crise tireotóxica."],

    # =========================================================================
    # DERMATOLÓGICOS INJECTÁVEIS
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 16. ADALIMUMAB (Psoríase, Crohn)
    # -------------------------------------------------------------------------
    ["Adalimumab", "", "", "SC", "dosagem(40, 80, 40, mg)", "40 mg/0.8ml|50", "14", "%[{#d/#c}] ml%#SC profunda. Caneta pré-carregada.#@@Infeções! TB latente!@", "OMS 2025. Psoríase, DII."],

    # =========================================================================
    # ANTIEMÉTICOS ADICIONAIS
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 17. DIMENIDRINATO (Náuseas, Vómitos)
    # -------------------------------------------------------------------------
    ["Dimenidrinato", "", "adulta", "IV", "peso(40, 120); dosagem(50, 50, 50, mg)", "50 mg/ml|50", "6, 8", "%[{#d/#c/#i}] ml%#Diluir em 10 ml SF e infundir em 5 min", "OMS 2025."],
    
    ["Dimenidrinato", "", "adulta", "IM", "peso(40, 120); dosagem(50, 50, 50, mg)", "50 mg/ml|50", "6, 8", "%[{#d/#c/#i}] ml%#IM profunda.", "OMS 2025."],

    # -------------------------------------------------------------------------
    # 18. DEXAMETASONA (Antiemético - quimioterapia)
    # -------------------------------------------------------------------------
    ["Dexametasona", "Náuseas", "adulta", "", "peso(40, 120); dosagem(4, 8, 8, mg)", "4 mg/ml|4; 8 mg/2ml|4", "8, 12", "%[{#d/#c/#i}] ml%#Diluir em 10 ml SF e infundir em 5 min", "OMS 2025. Antiemético QT."],

    # =========================================================================
    # RELAXANTES MUSCULARES
    # =========================================================================
    
    # -------------------------------------------------------------------------
    # 19. SUXAMETÓNIO (Succinilcolina) (Intubação rápida)
    # -------------------------------------------------------------------------
    ["Suxametónio", "", "adulta", "IV", "peso(40, 120); dosagem(1, 1.5, 1, mg/kg)", "50 mg/ml|50", "", "%[{#p*#d/#c}] ml%#IV directo rápido. Início: 30-60 seg.#@@Hipercalemia! Hipertermia maligna!@", "OMS 2025. MSF 2024. Sequência rápida."],
    
    ["Suxametónio", "", "pediatrica", "IV", "peso(3, 40); dosagem(1, 2, 2, mg/kg)", "50 mg/ml|50", "", "%[{#p*#d/#c}] ml%#IV directo rápido. Pré-medicar Atropina.", "OMS 2025."],
    
    ["Suxametónio", "", "adulta", "IM", "peso(40, 120); dosagem(3, 4, 3, mg/kg)", "50 mg/ml|50", "", "%[{#p*#d/#c}] ml%#IM profunda. Se IV impossível.", "OMS 2025."],

    # -------------------------------------------------------------------------
    # 20. VECURÓNIO (Bloqueio neuromuscular)
    # -------------------------------------------------------------------------
    ["Vecurónio", "", "adulta", "", "peso(40, 120); dosagem(0.08, 0.1, 0.1, mg/kg)", "10 mg/2ml|5", "", "%[{#p*#d/#c}] ml%#IV directo. Início: 2-3 min.#@@Paralisia prolongada!@", "OMS 2025. Intubação."],
    
    ["Vecurónio", "", "pediatrica", "", "peso(3, 40); dosagem(0.08, 0.1, 0.1, mg/kg)", "10 mg/2ml|5", "", "%[{#p*#d/#c}] ml%#IV directo.", "OMS 2025."],

    # -------------------------------------------------------------------------
    # 21. ROCURÓNIO (Bloqueio neuromuscular rápido)
    # -------------------------------------------------------------------------
    ["Rocurónio", "", "adulta", "", "peso(40, 120); dosagem(0.6, 1.2, 1, mg/kg)", "50 mg/5ml|10", "", "%[{#p*#d/#c}] ml%#IV directo. Início: 1-2 min.#@@Anafilaxia rara!@", "OMS 2025. Sequência rápida."],
    
    ["Rocurónio", "", "pediatrica", "", "peso(3, 40); dosagem(0.6, 1.2, 1, mg/kg)", "50 mg/5ml|10", "", "%[{#p*#d/#c}] ml%#IV directo.", "OMS 2025."],

    # -------------------------------------------------------------------------
    # 22. ATRACÚRIO (Bloqueio neuromuscular - UCI)
    # -------------------------------------------------------------------------
    ["Atracúrio", "", "adulta", "", "peso(40, 120); dosagem(0.3, 0.6, 0.5, mg/kg)", "50 mg/5ml|10", "", "%[{#p*#d/#c}] ml%#IV directo. Manutenção: infusão contínua.#@@Liberação histamina!@", "OMS 2025. UCI."],
    
    ["Atracúrio", "", "", "", "dosagem(5, 10, 5, mcg/kg/min)", "50 mg/5ml|10", "", "%[{#p*#d/#c}] ml/h%#Infusão contínua. Titular TOF.", "OMS 2025. UCI."],

]

motor_horizontal('medicamentos.xlsx', novos_dados)